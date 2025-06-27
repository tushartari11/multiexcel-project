import openpyxl
from collections import defaultdict
import json
import datetime
import pg_dbconnect
import re
import logging

def excel_to_dictionary(file_path):
    """
    Read multiple sheets from an Excel file and create dictionaries
    using headers as keys and cell values as values for each row.
    
    Args:
        file_path (str): Path to the Excel file
    
    Returns:
        dict: Dictionary with sheet names as keys and list of row dictionaries as values
    """
    
    # Open the Excel workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    # Dictionary to store all sheets data
    all_sheets_data = {}
    
    # Get all sheet names
    sheet_names = workbook.sheetnames
    print(f"Found {len(sheet_names)} sheets: {sheet_names}")
    logging.debug('Found %s sheets %s:',len(sheet_names), sheet_names)
    
    # Iterate through each sheet
    for sheet_name in sheet_names:
        logging.info("Processing sheet: %s", {sheet_name})
        
        # Select the current sheet
        worksheet = workbook[sheet_name]
        
        # Get the headers from the first row
        headers = []
        first_row = worksheet[1]  # First row (1-indexed)
        
        for cell in first_row:
            # Normalize header names
            normalized_header = normalize_headers(cell.value)
            if normalized_header is not None:
                headers.append(normalized_header.strip())
            else:
                headers.append(f"Column_{len(headers) + 1}")  # Default name for empty headers
        
        logging.info("Headers found: %s ", headers)
        logging.info("Total number of columns in header:  %s", len(headers))
        
        # List to store dictionaries for each row
        sheet_data = []
        
        # Iterate through rows starting from row 2 (skip header row)
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            # Create dictionary for current row
            row_dict = {}
            
            # Iterate through each cell in the row
            for col_index, cell_value in enumerate(row):
                # Make sure we don't exceed the number of headers
                if col_index < len(headers):
                    header = headers[col_index]
                    # Handle None values and convert to appropriate type
                    if cell_value is None:
                        row_dict[header] = ""
                    else:
                        row_dict[header] = cell_value
            
            # Add the row dictionary to sheet data
            sheet_data.append(row_dict)
           # print(f"Row {row_num}: {row_dict}")
        
        # Add sheet data to main dictionary
        all_sheets_data[sheet_name] = sheet_data
        logging.info('Sheet %s processed: %s rows', sheet_name, len(sheet_data))
    
    # Close the workbook
    workbook.close()
    
    return all_sheets_data

def normalize_headers(headerString):
    # Replace sequences of space and/or hyphen with single underscore
    if headerString is None:
        return None
    if not isinstance(headerString, str):
        headerString = str(headerString)
        logging.debug("In normalize_header : headerString:-  %s :: len:- %s", headerString, len(headerString))
    return re.sub(r'[\s\-]+', '_', headerString.strip())

def print_sheet_summary(data_dict):
    """
    Print a summary of the data structure
    """
    print("\n" + "="*50)
    print("DATA SUMMARY")
    print("="*50)
    

    for sheet_name, rows in data_dict.items():
        print(f"\nSheet: {sheet_name}")
        print(f"Number of rows: {len(rows)}")
        
        if rows:
            print(f"Columns: {list(rows[0].keys())}")
            print(f"Sample row: {rows[0]}")

# exports the dictionary to a Json file
def export_to_json(data_dict, output_file):
    """
    Export the data dictionary to a JSON file.
    
    Args:
        data_dict (dict): The data dictionary to export.
        output_file (str): The path to the output JSON file.
    """
    
    def json_serializer(obj):
        """JSON serializer for objects not serializable by default json code"""
        if isinstance(obj, (datetime.datetime, datetime.date)):
            return obj.isoformat()
        elif isinstance(obj, datetime.time):
            return obj.strftime('%H:%M:%S')
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
    
    with open(output_file, 'w', encoding='utf-8') as json_file:
        json.dump(data_dict, json_file, indent=4, ensure_ascii=False, default=json_serializer)
    print(f"Data exported to {output_file}")

# This method creates insert statements for each row in the dictionary
def create_insert_statements(data_dict, table_name):
    """
    Create SQL insert statements for each row in the data dictionary.
    
    Args:
        data_dict (dict): The data dictionary containing sheet data.
        table_name (str): The name of the database table to insert into.
    
    Returns:
        list: List of SQL insert statements.
    """
    insert_statements = []
    
    for sheet_name, rows in data_dict.items():
        for row in rows:
            columns = ', '.join(row.keys())
            values = ', '.join(f"'{str(value).replace('\'', '\'\'')}'" if value is not None else 'NULL' for value in row.values())
            insert_statement = f"INSERT INTO {table_name} ({columns}) VALUES ({values});"
            #print(f"executing statement: {insert_statement}")
            insert_statements.append(insert_statement)
    
    return insert_statements

# This is function to insert data into a PostgreSQL database
def insert_data_to_db(data_dict, table_name):
    """
    Insert data from the dictionary into a PostgreSQL database table.
    
    Args:
        data_dict (dict): The data dictionary containing sheet data.
        table_name (str): The name of the database table to insert into.
    
    Returns:
        bool: True if all data was inserted successfully, False otherwise.
    """
    conn = pg_dbconnect.create_connection()
    if not conn:
        logging.error("Failed to create a database connection.")
        return False
    
    cursor = conn.cursor()
    total_rows = 0
    successful_inserts = 0
    failed_inserts = 0
    
    try:
        # Clean the data before insertion
        cleaned_data = clean_data_for_insert(data_dict)
        
        for sheet_name, rows in cleaned_data.items():
            logging.info("Inserting data from sheet: %s (%s rows)", sheet_name, len(rows))
            
            for row_num, row in enumerate(rows, 1):
                total_rows += 1
                try:
                    # Only include columns that have values (non-empty)
                    # This allows database DEFAULT values to be used for omitted columns
                    if not row:  # Skip completely empty rows
                        logging.warning("Skipping empty row %s from sheet '%s'", row_num, sheet_name)
                        continue
                        
                    columns = ', '.join(row.keys())
                    placeholders = ', '.join(['%s'] * len(row))
                    insert_statement = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
                    
                    # Execute the insert with proper parameterization
                    cursor.execute(insert_statement, tuple(row.values()))
                    successful_inserts += 1
                    
                    logging.debug("Row %s from sheet '%s' inserted successfully", row_num, sheet_name)
                    
                except Exception as row_error:
                    failed_inserts += 1
                    logging.error("Failed to insert row %s from sheet '%s': %s", row_num, sheet_name, row_error)
                    logging.error("Row data: %s", row)
                    logging.error("Insert statement: %s", insert_statement)
                    # Continue with next row instead of stopping
                    continue
        
        # Commit all successful transactions
        conn.commit()
        logging.info("Data insertion completed. Total: %s, Successful: %s, Failed: %s", 
                    total_rows, successful_inserts, failed_inserts)
        
        if failed_inserts > 0:
            logging.warning("Some rows failed to insert. Check the log for details.")
            return False
        else:
            logging.info("All rows inserted successfully!")
            return True
            
    except Exception as e:
        logging.error("Critical error during data insertion: %s", e)
        logging.error("Rolling back all transactions...")
        conn.rollback()
        return False
        
    finally:
        cursor.close()
        logging.debug("Closing database connection.")
        pg_dbconnect.close_connection(conn)
        logging.debug("Database connection closed.")

# This function creates a CREATE TABLE statement based on the data structure
def create_table_statement(data_dict, table_name):
    """
    Generate a CREATE TABLE statement based on the data dictionary structure.
    
    Args:
        data_dict (dict): The data dictionary containing sheet data.
        table_name (str): The name of the database table to create.
    
    Returns:
        str: CREATE TABLE SQL statement.
    """
    if not data_dict:
        return None
    
    # Get the first sheet's first row to analyze column structure
    first_sheet = next(iter(data_dict.values()))
    if not first_sheet:
        return None
    
    first_row = first_sheet[0]
    
    # Analyze data types for each column
    columns = []
    columns.append("id SERIAL PRIMARY KEY")  # Add auto-increment primary key
    
    for column_name, sample_value in first_row.items():
        # Clean column name for SQL compatibility
        clean_column_name = column_name.replace(' ', '_').replace('-', '_').replace('.', '_').lower()
        
        # Determine data type based on sample value and add appropriate defaults
        if sample_value is None or sample_value == "":
            data_type = "TEXT DEFAULT ''"
        elif isinstance(sample_value, bool):
            data_type = "BOOLEAN DEFAULT FALSE"
        elif isinstance(sample_value, int):
            data_type = "INTEGER DEFAULT 0"
        elif isinstance(sample_value, float):
            data_type = "DECIMAL(10,2) DEFAULT 0.0"
        elif isinstance(sample_value, (datetime.datetime, datetime.date)):
            data_type = "DATE DEFAULT '1900-01-01'"
        elif isinstance(sample_value, datetime.time):
            data_type = "TIME DEFAULT '00:00:00'"
        else:
            # Default to TEXT with empty string default
            data_type = "TEXT DEFAULT ''"
        
        columns.append(f"{clean_column_name} {data_type}")
    
    columns_str = ",\n    ".join(columns)
    
    create_statement = f"""CREATE TABLE IF NOT EXISTS {table_name} (
    {columns_str}
);"""
    
    return create_statement

# This function executes the CREATE TABLE statement
def create_table_in_db(data_dict, table_name):
    """
    Create the table in the PostgreSQL database.
    
    Args:
        data_dict (dict): The data dictionary containing sheet data.
        table_name (str): The name of the database table to create.
    """
    conn = pg_dbconnect.create_connection()
    if not conn:
        logging.error("Failed to create a database connection.")
        return False
    
    cursor = conn.cursor()
    
    try:
        create_statement = create_table_statement(data_dict, table_name)
        if create_statement:
            logging.info("Executing CREATE TABLE statement:")
            logging.info(create_statement)
            cursor.execute(create_statement)
            conn.commit()
            logging.debug("Table %s created successfully!", table_name)
            return True
        else:
            logging.warning("Could not generate CREATE TABLE statement - no data found.")
            return False
    except Exception as e:
        logging.error("Error creating table: %s", e)
        logging.error("Rolling back the transaction due to error.")
        conn.rollback()
        return False
    finally:
        cursor.close()
        logging.debug("Closing database connection.")
        pg_dbconnect.close_connection(conn)
        logging.debug("Database connection closed.")

def clean_data_for_insert(data_dict):
    """
    Clean data for database insertion by handling empty values.
    Removes empty/None values so database can use DEFAULT values.
    
    Args:
        data_dict (dict): The data dictionary containing sheet data.
    
    Returns:
        dict: Cleaned data dictionary.
    """
    cleaned_data = {}
    
    for sheet_name, rows in data_dict.items():
        cleaned_rows = []
        
        for row in rows:
            cleaned_row = {}
            
            for column_name, value in row.items():
                # Handle empty or None values by omitting them from INSERT
                if value is not None and str(value).strip() != '':
                    cleaned_row[column_name] = value
                # If value is empty/None, skip it - let database use DEFAULT value
            
            cleaned_rows.append(cleaned_row)
        
        cleaned_data[sheet_name] = cleaned_rows
    
    return cleaned_data

def get_table_schema(data_dict):
    """
    Analyze the data structure and return a schema signature.
    
    Args:
        data_dict (dict): The data dictionary containing sheet data.
    
    Returns:
        tuple: A tuple representing the schema (column names and types).
    """
    if not data_dict:
        return None
    
    # Get the first sheet's first row to analyze column structure
    first_sheet = next(iter(data_dict.values()))
    if not first_sheet:
        return None
    
    first_row = first_sheet[0]
    
    # Create a schema signature based on column names and types
    schema_items = []
    for column_name, sample_value in first_row.items():
        # Clean column name for SQL compatibility
        clean_column_name = column_name.replace(' ', '_').replace('-', '_').replace('.', '_').lower()
        
        # Determine data type based on sample value
        if sample_value is None or sample_value == "":
            data_type = "TEXT"
        elif isinstance(sample_value, bool):
            data_type = "BOOLEAN"
        elif isinstance(sample_value, int):
            data_type = "INTEGER"
        elif isinstance(sample_value, float):
            data_type = "DECIMAL"
        elif isinstance(sample_value, (datetime.datetime, datetime.date)):
            data_type = "DATE"
        elif isinstance(sample_value, datetime.time):
            data_type = "TIME"
        else:
            data_type = "TEXT"
        
        schema_items.append((clean_column_name, data_type))
    
    return tuple(schema_items)

def compare_schemas(schema1, schema2):
    """
    Compare two table schemas to see if they're identical.
    
    Args:
        schema1 (tuple): First schema.
        schema2 (tuple): Second schema.
    
    Returns:
        bool: True if schemas are identical, False otherwise.
    """
    return schema1 == schema2

def group_sheets_by_schema(all_data):
    """
    Group sheets by their schema structure.
    
    Args:
        all_data (dict): Dictionary with all sheets data.
    
    Returns:
        dict: Dictionary where keys are schema signatures and values are sheet groups.
    """
    schema_groups = {}
    
    for sheet_name, sheet_data in all_data.items():
        if not sheet_data:  # Skip empty sheets
            logging.warning("Sheet '%s' is empty, skipping schema analysis", sheet_name)
            continue
            
        # Create a single-sheet dict for schema analysis
        single_sheet_dict = {sheet_name: sheet_data}
        schema = get_table_schema(single_sheet_dict)
        
        if schema:
            # Convert schema to a hashable string for grouping
            schema_key = str(schema)
            
            if schema_key not in schema_groups:
                schema_groups[schema_key] = {
                    'sheets': [],
                    'schema': schema,
                    'sample_data': single_sheet_dict
                }
            
            schema_groups[schema_key]['sheets'].append(sheet_name)
            # Add this sheet's data to the group
            if 'data' not in schema_groups[schema_key]:
                schema_groups[schema_key]['data'] = {}
            schema_groups[schema_key]['data'][sheet_name] = sheet_data
    
    return schema_groups

def create_tables_for_schema_groups(schema_groups, base_table_name):
    """
    Create separate tables for each schema group.
    
    Args:
        schema_groups (dict): Dictionary of schema groups.
        base_table_name (str): Base name for tables.
    
    Returns:
        dict: Dictionary mapping schema groups to table names.
    """
    table_mapping = {}
    
    for group_index, (schema_key, group_info) in enumerate(schema_groups.items(), 1):
        # Generate table name based on the sheets in this group
        sheets_in_group = group_info['sheets']
        
        # Determine table name based on sheet years/patterns
        if len(sheets_in_group) == 1:
            table_name = f"{base_table_name}_{sheets_in_group[0]}"
        else:
            # Try to find a pattern (like year ranges)
            years = []
            for sheet in sheets_in_group:
                if sheet.isdigit() and len(sheet) == 4:  # Year format
                    years.append(int(sheet))
            
            if years:
                years.sort()
                if len(years) > 1:
                    table_name = f"{base_table_name}_{min(years)}_to_{max(years)}"
                else:
                    table_name = f"{base_table_name}_{years[0]}"
            else:
                table_name = f"{base_table_name}_group_{group_index}"
        
        logging.info("Schema Group %s: Sheets %s -> Table '%s'", group_index, sheets_in_group, table_name)
        logging.info("Columns in this schema: %s", [col[0] for col in group_info['schema']])
        
        # Create table for this schema group
        if create_table_in_db(group_info['sample_data'], table_name):
            table_mapping[schema_key] = table_name
            logging.info("Table '%s' created successfully for sheets: %s", table_name, sheets_in_group)
        else:
            logging.error("Failed to create table '%s' for sheets: %s", table_name, sheets_in_group)
    
    return table_mapping

# Example usage
if __name__ == "__main__":

    logging.basicConfig(filename='excel_to_dict.log', level=logging.DEBUG, filemode='w', format='%(asctime)s - %(levelname)s - %(message)s')

    # Replace with your Excel file path
    excel_file_path = "/Users/tushartari/tushar/study/courses/IraSkills/work/JCB_DATA_PUNE_CLEANED.xlsx"

    output_json_file = "/Users/tushartari/tushar/study/courses/IraSkills/work/JCB_DATA_PUNE_CLEANED.json"
    
    try:
        # Read the Excel file and convert to dictionary
        result = excel_to_dictionary(excel_file_path)
        
        # Print summary
        #print_sheet_summary(result)
        # Export to JSON file
        #export_to_json(result, output_json_file)
        
        # Group sheets by schema structure
        logging.info("Analyzing sheet structures...")
        schema_groups = group_sheets_by_schema(result)
        
        logging.info("Found %s different schema groups:", len(schema_groups))
        for i, (schema_key, group_info) in enumerate(schema_groups.items(), 1):
            logging.info("Group %s: Sheets %s", i, group_info['sheets'])
        
        # Create tables for each schema group
        base_table_name = "invoice_data"
        table_mapping = create_tables_for_schema_groups(schema_groups, base_table_name)
        
        # Insert data into appropriate tables
        for schema_key, group_info in schema_groups.items():
            if schema_key in table_mapping:
                table_name = table_mapping[schema_key]
                group_data = group_info['data']
                
                logging.info("Inserting data for schema group with sheets: %s into table '%s'", 
                           group_info['sheets'], table_name)
                
                if insert_data_to_db(group_data, table_name):
                    logging.info("Data insertion completed successfully for table '%s'!", table_name)
                else:
                    logging.error("Data insertion completed with errors for table '%s'. Check the log for details.", table_name)
            else:
                logging.error("No table created for schema group with sheets: %s", group_info['sheets'])
    except Exception as e:
        print(f"An error occurred: {e}")