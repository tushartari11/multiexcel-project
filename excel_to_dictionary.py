import openpyxl
from collections import defaultdict
import json
import datetime
import re
import logging

def excel_to_dictionary(file_path):
    """
    Read multiple sheets from an Excel file and create dictionaries
    using headers as keys and cell values as values for each row.
    
    Args:
        file_path (str): Path to the Excel file
    
    Returns:
        dict: Dictionary with sheet names as keys and list of row dictionaries as values
    """
    
    # Open the Excel workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    # Dictionary to store all sheets data
    all_sheets_data = {}
    
    # Get all sheet names
    sheet_names = workbook.sheetnames
    print(f"Found {len(sheet_names)} sheets: {sheet_names}")
    logging.debug('Found %s sheets %s:',len(sheet_names), sheet_names)
    
    # Iterate through each sheet
    for sheet_name in sheet_names:
        logging.info("Processing sheet: %s", {sheet_name})
        
        # Select the current sheet
        worksheet = workbook[sheet_name]
        
        # Get the headers from the first row
        headers = []
        first_row = worksheet[1]  # First row (1-indexed)
        
        for cell in first_row:
            # Normalize header names
            normalized_header = normalize_headers(cell.value)
            if normalized_header is not None:
                headers.append(normalized_header.strip())
            else:
                headers.append(f"Column_{len(headers) + 1}")  # Default name for empty headers
        
        logging.info("Headers found: %s ", headers)
        logging.info("Total number of columns in header:  %s", len(headers))
        
        # List to store dictionaries for each row
        sheet_data = []
        
        # Iterate through rows starting from row 2 (skip header row)
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            # Create dictionary for current row
            row_dict = {}
            
            # Iterate through each cell in the row
            for col_index, cell_value in enumerate(row):
                # Make sure we don't exceed the number of headers
                if col_index < len(headers):
                    header = headers[col_index]
                    # Handle None values and convert to appropriate type
                    if cell_value is None:
                        row_dict[header] = ""
                    else:
                        row_dict[header] = cell_value
            
            # Add the row dictionary to sheet data
            sheet_data.append(row_dict)
           # print(f"Row {row_num}: {row_dict}")
        
        # Add sheet data to main dictionary
        all_sheets_data[sheet_name] = sheet_data
        logging.info('Sheet %s processed: %s rows', sheet_name, len(sheet_data))
    
    # Close the workbook
    workbook.close()
    
    return all_sheets_data


def normalize_headers(headerString):
    # Replace sequences of space and/or hyphen with single underscore
    if headerString is None:
        return None
    if not isinstance(headerString, str):
        headerString = str(headerString)
        logging.debug("In normalize_header : headerString:-  %s :: len:- %s", headerString, len(headerString))
    return re.sub(r'[\s\-]+', '_', headerString.strip())