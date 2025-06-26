from openpyxl import load_workbook
import csv

workbook = load_workbook('/Users/tushartari/tushar/study/courses/IraSkills/work/JCB_DATA_PUNE_CLEANED.xlsx',data_only=True)
sheet_names = workbook.sheetnames

print(f"Total number of sheets: {len(sheet_names)}")
for sheet_name in sheet_names:
    print(f"Processing sheet: {sheet_name}")
    worksheet = workbook[sheet_name]
    
    csv_file_path = f"{sheet_name}.csv"
    with open(csv_file_path, 'w', newline='', encoding='utf-8') as csv_file:
        csv_writer = csv.writer(csv_file)
    
        for row in worksheet.iter_rows(values_only=True):
            csv_writer.writerow(row)
    
    print(f"Data from sheet '{sheet_name}' has been written to {csv_file_path}")